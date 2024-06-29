from django.http import HttpResponse, HttpResponseRedirect
from django.shortcuts import get_object_or_404, render
from django.urls import reverse
from .models import *
from django.contrib import messages
from django.contrib.auth.models import User
from django.contrib.auth import authenticate, login, logout
from django.shortcuts import redirect
from django.contrib.auth.decorators import login_required, permission_required
from .forms import *
from django.core.paginator import Paginator
from django.db import IntegrityError
from tablib import Dataset
from .resources import *
from django.contrib.auth.forms import UserCreationForm, UserChangeForm
from openpyxl import Workbook
import pytz
from reportlab.lib.pagesizes import letter
from reportlab.lib.units import inch
from reportlab.pdfgen import canvas
from reportlab.platypus import Image, Spacer, Paragraph, Table, TableStyle, SimpleDocTemplate
from reportlab.lib.styles import getSampleStyleSheet
from reportlab.lib import colors
from reportlab.lib.utils import ImageReader
from django.core.files.storage import default_storage
from reportlab.platypus import Image
from reportlab.lib.colors import white

from io import BytesIO
import qrcode
from PIL import Image
# Create your views here.    
def login_user(request):
    if request.method == "POST":
        username = request.POST['username']
        password = request.POST['password']
        user = authenticate(request, username=username, password=password)
        if user is not None:
            login(request, user)
            messages.success(request, "Inicio de sesión exitoso. ¡Bienvenido!")
            return redirect('index')
          
        else:
            messages.error(request, "Credenciales inválidas. Por favor, verifica tus credenciales e inténtalo nuevamente.")
            return redirect('login')

    return render(request, 'login/login.html', {})

def logout_user(request):
    logout(request)
    messages.success(request, ("Sesion Cerrada Exitosamente"))
    return redirect('login')

@login_required
def index(request):
    
    return render(request,'index2.html', {

    })

@login_required
def setting(request):
    return render(request,'settings.html',{ })

@login_required
def citas(request):
    citas = Citas.objects.all().order_by('paciente')
    paginator = Paginator(citas, 6)  # mostrar 10 citas por página
    page = request.GET.get('page')
    page_obj = paginator.get_page(page)
    return render(request, 'citas/citas.html', {'cita': page_obj})

@login_required
def tickets(request):
    tickets = Ticket.objects.all().order_by('id')
    paginator = Paginator(tickets, 6)  # mostrar 10 citas por página
    page = request.GET.get('page')
    page_obj = paginator.get_page(page)
    return render(request, 'tickets/ticket.html', {'ticket': page_obj})

@login_required
def medicos(request):
    if request.method == 'POST':
        form = ImportExcelForm(request.POST, request.FILES)
        if form.is_valid():
            medicos_resource = MedicosResource()
            dataset = Dataset()
            new_medicos = request.FILES['my_file']
            imported_data = dataset.load(new_medicos.read(), format='xlsx')
            try:
                for data in imported_data:
                    value = Medicos(
                        data[0],
                        data[1],
                        data[2],
                        data[3],
                        data[4],
                        data[5],
                        data[6]

                    )
                    value.save()
            except IntegrityError:
                messages.warning(request, "Algunos registros ya existían y no fueron cargados.")

            else:
                messages.success(request, "Carga masiva realizada con éxito.")

            return redirect('medicos')
    else:
        form = ImportExcelForm()
    medicos = Medicos.objects.all().order_by('name')
    paginator = Paginator(medicos, 6)  # muestra 10 pacientes por página
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)
    context = {
        'medico': page_obj
    }
    return render(request, 'medicos/medicos.html', context)

@login_required
def usuarios(request):
    users = User.objects.all()
    return render(request, 'usuarios/usuarios.html', {
        'users': users
    })

@login_required
def edit_profile(request):
    if request.method == 'POST':
        form = CustomUserChangeForm(request.POST, instance=request.user)
        if form.is_valid():
            form.save()
    else:
        form = CustomUserChangeForm(instance=request.user)
    return render(request, 'usuarios/editUser.html', {'form': form})

@login_required
def pacientes(request):
    if request.method == 'POST':
        form = ImportExcelForm(request.POST, request.FILES)
        if form.is_valid():
            pacientes_resource = PacientesResource()
            dataset = Dataset()
            new_pacientes = request.FILES['my_file']
            imported_data = dataset.load(new_pacientes.read(), format='xlsx')
            
            for data in imported_data:
                    value = Pacientes(
                        data[0],
                        data[1],
                        data[2],
                        data[3],
                        data[4],
                        data[5],
                        data[6],
                        data[7]

                    )
                    value.save()
           

           
            messages.success(request, "Carga masiva realizada con éxito.")

            return redirect('pacientes')
    else:
        form = ImportExcelForm()
    pacientes = Pacientes.objects.all().order_by('name')
    paginator = Paginator(pacientes, 6)  # muestra 10 pacientes por página
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)
    context = {
        'paciente': page_obj
    }
    return render(request, 'pacientes/pacientes.html', context)

@login_required
def add(request):
    if request.method == 'POST':
        form = PacientesForm(request.POST)
        if form.is_valid():
            new_Nombre_Paciente= form.cleaned_data['name']
            new_LastName_Paciente= form.cleaned_data['lastName']
            new_Email=form.cleaned_data['email']
            new_Dni=form.cleaned_data['dni']
            new_Cel=form.cleaned_data['celp']
            new_Address=form.cleaned_data['address']
            new_FechaNac = form.cleaned_data['fechaNac']
            new_seguro=form.cleaned_data['seguro']

            new_Paciente =Pacientes(
                name = new_Nombre_Paciente,
                lastName = new_LastName_Paciente,
                email = new_Email,
                dni = new_Dni,
                celp = new_Cel,
                address = new_Address,
                fechaNac=new_FechaNac,
                seguro = new_seguro
            )
            new_Paciente.save()
            return render(request,'pacientes/add.html',{
                          'form': PacientesForm(),
                            'success': True
            })
        else:
            form =PacientesForm()
        return  render(request,'pacientes/add.html',
                       {'form' : PacientesForm() })
    else:
        form = PacientesForm()
        return render(request, 'pacientes/add.html', {'form': form})

@login_required
def edit(request, id):
    paciente = Pacientes.objects.get(pk=id)
    if request.method == 'POST':
        form = PacientesForm(request.POST, instance=paciente)
        if form.is_valid():
            form.save()
            return render(request, 'pacientes/edit.html',{
                'form': form,
                'success': True
            })
        else:
            # en caso de que el formulario no sea válido, renderiza el template con el formulario pre-cargado con los datos del paciente existente
            return render(request, 'pacientes/edit.html', {
                'form': form,
                'paciente': paciente
            })
    else:
        # en caso de una solicitud GET, renderiza el formulario pre-cargado con los datos del paciente existente
        form = PacientesForm(instance=paciente)
        return render(request, 'pacientes/edit.html', {'form': form, 'paciente': paciente})

@login_required
def delete(request, id):
    if request.method == 'POST':
        paciente = Pacientes.objects.get(pk=id)
        paciente.delete()
    return HttpResponseRedirect(reverse('pacientes'))

@login_required
def Export_Pacientes(request):
    citas = Pacientes.objects.all()

    # Crear archivo Excel
    wb = Workbook()
    ws = wb.active

    # Añadir encabezados de columnas
    ws['A1'] = 'Paciente'
    ws['B1'] = 'Dni '
    ws['C1'] = 'Telefono'
    ws['D1'] = 'Email'
    ws['E1'] = 'Direccion'
    ws['F1'] = 'Fecha Nacimiento'
    ws['G1'] = 'Seguro'

    # Añadir datos de las citas a las filas
    row_num = 2
    for cita in citas:
        ws.cell(row=row_num, column=1, value=cita.name + ' ' + cita.lastName)
        ws.cell(row=row_num, column=2, value=cita.dni)
        ws.cell(row=row_num, column=3, value=cita.celp)
        tz = pytz.timezone('America/Bogota')  # o la zona horaria que corresponda
        ws.cell(row=row_num, column=4, value=cita.email)
        ws.cell(row=row_num, column=5, value=cita.address)
        ws.cell(row=row_num, column=6, value=cita.fechaNac)
        ws.cell(row=row_num, column=7, value=cita.seguro.nombre)
        row_num += 1

    # Establecer el nombre del archivo Excel
    filename = 'Pacientes.xlsx'

    # Devolver el archivo Excel como una respuesta de descarga
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename=' + filename
    wb.save(response)
    return response

@login_required
def create_user(request):
    if request.method == 'POST':
        form = CustomUserCreationForm(request.POST)
        if form.is_valid():
            user = form.save()
            return HttpResponseRedirect(reverse('usuarios'))
    else:
        form = CustomUserCreationForm()
    return render(request, 'usuarios/addUsuarios.html', {'form': form})

@login_required
def editUsuarios(request, user_id):
    user = get_object_or_404(User, pk=user_id)
    if request.method == 'POST':
        form = CustomUserChangeForm(request.POST, instance=user)
        if form.is_valid():
            form.save()
            return redirect('usuarios')
    else:
        form = CustomUserChangeForm(instance=user)
    return render(request, 'usuarios/editUsuarios.html', {'form': form})

@login_required
def deleteUsuarios(request, user_id):
    if request.method == 'POST':
        user = get_object_or_404(User, pk=user_id)
        user.delete()
    return HttpResponseRedirect(reverse('usuarios'))

@login_required
def addMedico(request):
    if request.method == 'POST':
        form = MedicosForm(request.POST)
        if form.is_valid():
            new_Nombre_Medico= form.cleaned_data['name']
            new_LastName_Medico= form.cleaned_data['lastName']
            new_Email=form.cleaned_data['email']
            new_Dni=form.cleaned_data['dni']
            new_Cel=form.cleaned_data['celp']
            new_Tipo=form.cleaned_data['tipo']

            new_Medico =Medicos(
                name = new_Nombre_Medico,
                lastName = new_LastName_Medico,
                email = new_Email,
                dni = new_Dni,
                celp = new_Cel,
                tipo = new_Tipo,
            )
            new_Medico.save()
            return render(request,'medicos/addMedico.html',{
                          'form': MedicosForm(),
                            'success': True
            })
        else:
            form =MedicosForm()
        return  render(request,'medicos/addMedico.html',
                       {'form' : MedicosForm() })
    else:
        form = MedicosForm()
        return render(request, 'medicos/addMedico.html', {'form': form})

@login_required
def editMedico(request, id):
    medico = Medicos.objects.get(pk=id)
    if request.method == 'POST':
        form = MedicosForm(request.POST, instance=medico)
        if form.is_valid():
            form.save()
            return render(request, 'medicos/editMedico.html',{
                'form': form,
                'success': True
            })
        else:
            # en caso de que el formulario no sea válido, renderiza el template con el formulario pre-cargado con los datos del paciente existente
            return render(request, 'medicos/editMedico.html', {
                'form': form,
                'medico': medico
            })
    else:
        # en caso de una solicitud GET, renderiza el formulario pre-cargado con los datos del paciente existente
        form = MedicosForm(instance=medico)
        return render(request, 'medicos/editMedico.html', {'form': form, 'medico': medico})

@login_required
def deleteMedico(request, id):
    if request.method == 'POST':
        medico = Medicos.objects.get(pk=id)
        medico.delete()
    return HttpResponseRedirect(reverse('medicos'))

@login_required
def Export_medicos(request):
    citas = Medicos.objects.all()

    # Crear archivo Excel
    wb = Workbook()
    ws = wb.active

    # Añadir encabezados de columnas
    ws['A1'] = 'Medico'
    ws['B1'] = 'Dni '
    ws['C1'] = 'Telefono'
    ws['D1'] = 'Email'
    ws['E1'] = 'Especialidad'

    # Añadir datos de las citas a las filas
    row_num = 2
    for cita in citas:
        ws.cell(row=row_num, column=1, value=cita.name + ' ' + cita.lastName)
        ws.cell(row=row_num, column=2, value=cita.dni)
        ws.cell(row=row_num, column=3, value=cita.celp)
        tz = pytz.timezone('America/Bogota')  # o la zona horaria que corresponda
        ws.cell(row=row_num, column=4, value=cita.email)
        ws.cell(row=row_num, column=5, value=cita.tipo.nombre)
        row_num += 1

    # Establecer el nombre del archivo Excel
    filename = 'Medicos.xlsx'

    # Devolver el archivo Excel como una respuesta de descarga
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename=' + filename
    wb.save(response)
    return response

@login_required
def especialidadList(request):
    pacientes = Especialidad.objects.all().order_by('nombre')
    paginator = Paginator(pacientes, 6)  # muestra 10 pacientes por página
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)
    context = {
        'paciente': page_obj
    }
    return render(request,'especialidad/especialidad.html',context)
    
    
@login_required
def createEspecialidad(request):
    if request.method == 'POST':
        form = EspecialidadForm(request.POST)
        if form.is_valid():
            new_Nombre_Especialidad= form.cleaned_data['nombre']

            new_Especialidad =Especialidad(
                nombre = new_Nombre_Especialidad,
            )
            new_Especialidad.save()
            return render(request,'especialidad/addEspecialidad.html',{
                          'form': EspecialidadForm(),
                            'success': True
            })
        else:
            form =EspecialidadForm()
        return  render(request,'especialidad/addEspecialidad.html',
                       {'form' : EspecialidadForm() })
    else:
        form = EspecialidadForm()
        return render(request, 'especialidad/addEspecialidad.html', {'form': form})
    
@login_required
def editarEspecialidad(request,id):
    especialidaddes = Especialidad.objects.get(pk=id)
    if request.method == 'POST':
        form = EspecialidadForm(request.POST, instance=especialidaddes)
        if form.is_valid():
            form.save()
            return render(request, 'especialidad/editEspecialidad.html',{
                'form': form,
                'success': True
            })
        else:
            # en caso de que el formulario no sea válido, renderiza el template con el formulario pre-cargado con los datos del paciente existente
            return render(request, 'especialidad/editEspecialidad.html', {
                'form': form,
                'paciente': especialidaddes
            })
    else:
        # en caso de una solicitud GET, renderiza el formulario pre-cargado con los datos del paciente existente
        form = EspecialidadForm(instance=especialidaddes)
        return render(request, 'especialidad/editEspecialidad.html', {'form': form, 'paciente': especialidaddes})
  
@login_required    
def deleteEspecialidad(request,id):
    if request.method == 'POST':
        especialidades2 = Especialidad.objects.get(pk=id)
        especialidades2.delete()
    return HttpResponseRedirect(reverse('especialidades'))

@login_required
def crear_cita(request):
    if request.method == 'POST':
        form = CitaForm(request.POST)
        if form.is_valid():
            nueva_cita = form.save(commit=False)
            nueva_cita.medico = request.user
            nueva_cita.save()
            return redirect('citas')
    else:
        form = CitaForm()

    context = {'form': form}
    return render(request, 'citas/addCita.html', context)

@login_required
def editar_cita(request, id):
    # Obtener la cita que se quiere editar
    cita = get_object_or_404(Citas, pk=id)

    if request.method == 'POST':
        # Si se envió el formulario con datos, procesarlos
        form = CitaForm(request.POST, instance=cita)
        if form.is_valid():
            form.save()
            return redirect('citas')
    else:
        # Si se accede a la vista por primera vez, mostrar el formulario con los datos actuales de la cita
        form = CitaForm(instance=cita)

    return render(request, 'citas/editCita.html', {'form': form, 'cita': cita})

@login_required
def delete_Citas(request, id):
    if request.method == 'POST':
        cita = get_object_or_404(Citas, pk=id)
        cita.delete()
    return HttpResponseRedirect(reverse('citas'))

@login_required
def Export_citas(request):
    citas = Citas.objects.all()

    # Crear archivo Excel
    wb = Workbook()
    ws = wb.active

    # Añadir encabezados de columnas
    ws['A1'] = 'Paciente'
    ws['B1'] = 'Dni Paciente'
    ws['C1'] = 'Telefono'
    ws['D1'] = 'Fecha Emision'
    ws['E1'] = 'Fecha Programada'
    ws['F1'] = 'Motivo'

    # Añadir datos de las citas a las filas
    row_num = 2
    for cita in citas:
        ws.cell(row=row_num, column=1, value=cita.paciente.name + ' ' + cita.paciente.lastName)
        ws.cell(row=row_num, column=2, value=cita.paciente.dni)
        ws.cell(row=row_num, column=3, value=cita.paciente.celp)
        tz = pytz.timezone('America/Bogota')  # o la zona horaria que corresponda
        ws.cell(row=row_num, column=4, value=cita.fechaEmision.astimezone(tz).replace(tzinfo=None))
        ws.cell(row=row_num, column=5, value=cita.fechaProgramada.astimezone(tz).replace(tzinfo=None))
        ws.cell(row=row_num, column=6, value=cita.motivo)
        row_num += 1

    # Establecer el nombre del archivo Excel
    filename = 'citas.xlsx'

    # Devolver el archivo Excel como una respuesta de descarga
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename=' + filename
    wb.save(response)
    return response

@login_required
def generar_ticket(request):
    if request.method == 'POST':
        form = TicketForm(request.POST)
        if form.is_valid():
            ticket = form.save()
            # Redireccionar al usuario a la página de confirmación de ticket
            return redirect('tickets')
    else:
        form = TicketForm()
    return render(request, 'tickets/addticket.html', {'form': form})

@login_required
def editar_ticket(request, id):
    # Obtener la cita que se quiere editar
    ticket = get_object_or_404(Ticket, pk=id)

    if request.method == 'POST':
        # Si se envió el formulario con datos, procesarlos
        form = TicketForm(request.POST, instance=ticket)
        if form.is_valid():
            form.save()
            return redirect('tickets')
    else:
        # Si se accede a la vista por primera vez, mostrar el formulario con los datos actuales de la cita
        form = TicketForm(instance=ticket)

    return render(request, 'tickets/editTicket.html', {'form': form, 'ticket': ticket})

@login_required
def delete_Ticket(request, id):
    if request.method == 'POST':
        ticket = get_object_or_404(Ticket, pk=id)
        ticket.delete()
    return HttpResponseRedirect(reverse('tickets'))

@login_required
def TicketPDFView(request, pk):
    width, height = letter
    response = HttpResponse(content_type='application/pdf')

    ticket = Ticket.objects.get(pk=pk)
    num_ticket = ticket.nroTicket

    response['Content-Disposition'] = f'attachment; filename="Ticket_{num_ticket}.pdf'

    buffer = BytesIO()
    c = canvas.Canvas(buffer, pagesize=letter)

    # Añadir imagen en la parte superior derecha
    logo_path = 'integrador/static/img/logopdf2.png'
    x = 2.5 * inch
    y = 9 * inch
    width = 4 * inch
    height = 2 * inch
    c.drawImage(logo_path, x, y, width, height)


    # Añadir título
    c.setFont('Helvetica-Bold', 20)
    c.drawCentredString(4.25*inch, 8.5*inch, 'Ticket de Atencion')

    c.drawString(0, 7.5 * inch, '')

    # Añadir tabla de ticket
    data = [
        ['Paciente:', ticket.paciente.name + " " + ticket.paciente.lastName],
        ['Código de Ticket:', ticket.nroTicket],
        ['Fecha Emitida:', ticket.fechaTicket.strftime('%d-%m-%Y %H:%M:%S')],
        ['Estado:', ticket.estado],
    ]
    table = Table(data, colWidths=[2*inch, 4*inch])

    table.setStyle(TableStyle([
        ('FONT', (0, 0), (-1, -1), 'Helvetica'),
        ('FONTSIZE', (0, 0), (-1, -1), 12),
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
        ('BACKGROUND', (0, 0), (-1, 0), '#f0f0f0'),
        ('TEXTCOLOR', (0, 0), (-1, 0), 'black'),
        ('GRID', (0, 0), (-1, -1), 1, 'black'),
        ('VALIGN', (0, 0), (-1, -1), 'MIDDLE')
    ]))

    # Obtener el ancho y alto de la tabla
    width, height = table.wrapOn(c, 0, 0)

    # Centrar la tabla en la página
    x = (letter[0] - width) / 2
    y = 6.5 * inch
    table.drawOn(c, x, y)

    # Crear los datos que deseas codificar en el QR
    qr_data = f'Paciente: {ticket.paciente.name} {ticket.paciente.lastName}\n' \
              f'Código de Ticket: {ticket.nroTicket}\n' \
              f'Fecha Emitida: {ticket.fechaTicket.strftime("%d-%m-%Y %H:%M:%S")}\n' \
              f'Estado: {ticket.estado}'

    # Crear un objeto QRCode
    qr = qrcode.QRCode(
        version=1,
        error_correction=qrcode.constants.ERROR_CORRECT_L,
        box_size=10,
        border=4,
    )

    # Agregar los datos al QRCode
    qr.add_data(qr_data)
    qr.make(fit=True)

    # Crear una imagen QRCode
    qr_image = qr.make_image(fill_color="black", back_color="white")
    qr_buffer = BytesIO()
    qr_image.save(qr_buffer, "PNG")

    # Crear un objeto ImageReader para agregar la imagen QRCode al PDF
    qr_image_reader = ImageReader(qr_buffer)
    
    # Dibujar el código QR en el PDF
    c.drawImage(qr_image_reader, 0.5*inch, 0.5*inch, width=1.5*inch, height=1.5*inch)


    c.save()

    pdf = buffer.getvalue()
    buffer.close()
    response.write(pdf)

    return response

@login_required
def Export_Tickets(request):
    tickets = Ticket.objects.all()

    # Crear archivo Excel
    wb = Workbook()
    ws = wb.active

    # Añadir encabezados de columnas
    ws['A1'] = 'Nro Ticket'
    ws['B1'] = 'Paciente'
    ws['C1'] = 'Dni Paciente'
    ws['D1'] = 'Telefono'
    ws['E1'] = 'Fecha Ticket'
    ws['F1'] = 'Estado'

    # Añadir datos de las citas a las filas
    row_num = 2
    for ticket in tickets:
        ws.cell(row=row_num, column=1, value=ticket.nroTicket)
        ws.cell(row=row_num, column=2, value=ticket.paciente.name + ' ' + ticket.paciente.lastName)
        ws.cell(row=row_num, column=3, value=ticket.paciente.dni)
        ws.cell(row=row_num, column=4, value=ticket.paciente.celp)
        tz = pytz.timezone('America/Bogota')  # o la zona horaria que corresponda
        ws.cell(row=row_num, column=5, value=ticket.fechaTicket.astimezone(tz).replace(tzinfo=None))
        ws.cell(row=row_num, column=6, value=ticket.estado)
        row_num += 1

    # Establecer el nombre del archivo Excel
    filename = 'tickets.xlsx'

    # Devolver el archivo Excel como una respuesta de descarga
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename=' + filename
    wb.save(response)
    return response

@login_required
def historias(request):
    historias = Historias.objects.all().order_by('paciente')
    paginator = Paginator(historias, 6)  # mostrar 10 citas por página
    page = request.GET.get('page')
    page_obj = paginator.get_page(page)
    return render(request, 'historias/historias.html', {'historia': page_obj})

@login_required
def Export_Historias(request):
    historias = Historias.objects.all()

    # Crear archivo Excel
    wb = Workbook()
    ws = wb.active

    # Añadir encabezados de columnas
    ws['A1'] = 'Paciente'
    ws['B1'] = 'DNI'
    ws['C1'] = 'Medico'
    ws['D1'] = 'Cita'
    ws['E1'] = 'Observaciones'


    # Añadir datos de las citas a las filas
    row_num = 2
    for historia in historias:
        ws.cell(row=row_num, column=1, value=historia.paciente.name + ' ' + historia.paciente.lastName)
        ws.cell(row=row_num, column=2, value=historia.paciente.dni)
        ws.cell(row=row_num, column=3, value=historia.medico.name+' '+historia.medico.lastName)
        tz = pytz.timezone('America/Bogota')  # o la zona horaria que corresponda
        ws.cell(row=row_num, column=4, value=historia.cita.fechaProgramada.astimezone(tz).replace(tzinfo=None))
        ws.cell(row=row_num, column=5, value=historia.observacion)
        row_num += 1

    # Establecer el nombre del archivo Excel
    filename = 'Historias.xlsx'

    # Devolver el archivo Excel como una respuesta de descarga
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename=' + filename
    wb.save(response)
    return response

@login_required
def HistoriaPDFView(request, pk):
    # Obtener la historia
    historia = get_object_or_404(Historias, pk=pk)

    # Crear un objeto PDF en memoria
    buffer = BytesIO()
    c = canvas.Canvas(buffer, pagesize=letter)
    # Crear el documento PDF con ReportLab
    doc = SimpleDocTemplate(buffer, pagesize=letter)
    styles = getSampleStyleSheet()
    style = styles['Normal']

    # Añadir imagen en la parte superior derecha
    logo_path = 'integrador/static/img/logopdf2.png'
    x = 2.5 * inch
    y = 9 * inch
    width = 4 * inch
    height = 2 * inch
    c.drawImage(logo_path, x, y, width, height)
    

    # Añadir título
    
    c.setFont('Helvetica-Bold', 20)
    c.drawCentredString(4.25*inch, 8.5*inch, 'Historia Clinica')

    c.drawString(0, 7.5 * inch, '')

    # Crear tabla con datos de la historia
    data = [
        ['Paciente:', f'{historia.paciente.name} {historia.paciente.lastName}'],
        ['Médico:', f'{historia.medico.name} {historia.medico.lastName}'],
        ['Fecha:', f'{historia.cita.fechaProgramada}'],
        ['Observación:', f'{historia.observacion}'],
    ]
    t = Table(data, colWidths=[2*inch, 4*inch])
    t.setStyle(TableStyle([
        ('FONT', (0, 0), (-1, -1), 'Helvetica'),
        ('FONTSIZE', (0, 0), (-1, -1), 12),
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
        ('BACKGROUND', (0, 0), (-1, 0), '#f0f0f0'),
        ('TEXTCOLOR', (0, 0), (-1, 0), 'black'),
        ('GRID', (0, 0), (-1, -1), 1, 'black'),
        ('VALIGN', (0, 0), (-1, -1), 'MIDDLE')
    ]))
    print(t)
      # Obtener el ancho y alto de la tabla
    width, height = t.wrapOn(c, 0, 0)

    # Centrar la tabla en la página
    x = (letter[0] - width) / 2
    y = 6.5 * inch
    t.drawOn(c, x, y)
   
 

    # Configurar la respuesta HTTP
    response = HttpResponse(content_type='application/pdf')
    response['Content-Disposition'] = f'inline; filename="HistoriaClinica_{historia.id}.pdf"'

     # Crear los datos que deseas codificar en el QR
    qr_data = f'Paciente: {historia.paciente.name} {historia.paciente.lastName}\n' \
              f'Médico: {historia.medico.name} {historia.medico.lastName}\n' \
              f'Fecha : {historia.cita.fechaProgramada}\n' \
              f'Observación: {historia.observacion}'

    # Crear un objeto QRCode
    qr = qrcode.QRCode(
        version=1,
        error_correction=qrcode.constants.ERROR_CORRECT_L,
        box_size=10,
        border=4,
    )

    # Agregar los datos al QRCode
    qr.add_data(qr_data)
    qr.make(fit=True)

    # Crear una imagen QRCode
    qr_image = qr.make_image(fill_color="black", back_color="white")
    qr_buffer = BytesIO()
    qr_image.save(qr_buffer, "PNG")

    # Crear un objeto ImageReader para agregar la imagen QRCode al PDF
    qr_image_reader = ImageReader(qr_buffer)
    
    # Dibujar el código QR en el PDF
    c.drawImage(qr_image_reader, 0.5*inch, 0.5*inch, width=1.5*inch, height=1.5*inch)


    c.save()
    # Adjuntar el PDF generado a la respuesta HTTP
    response.write(buffer.getvalue())
    buffer.close()

    return response

@login_required
def generar_historia(request):
    if request.method == 'POST':
        form = HistoriasForm(request.POST)
        if form.is_valid():
            ticket = form.save()
            # Redireccionar al usuario a la página de confirmación de ticket
            return redirect('historias')
    else:
        form = HistoriasForm()
    return render(request, 'historias/addHistoria.html', {'form': form})

@login_required
def editar_historia(request, id):
    # Obtener la cita que se quiere editar
    historia = get_object_or_404(Historias, pk=id)

    if request.method == 'POST':
        # Si se envió el formulario con datos, procesarlos
        form = HistoriasForm(request.POST, instance=historia)
        if form.is_valid():
            form.save()
            return redirect('historias')
    else:
        # Si se accede a la vista por primera vez, mostrar el formulario con los datos actuales de la cita
        form = HistoriasForm(instance=historia)

    return render(request, 'historias/editHistoria.html', {'form': form, 'historia': historia})

@login_required
def delete_historia(request, id):
    if request.method == 'POST':
        historia = get_object_or_404(Historias, pk=id)
        historia.delete()
    return HttpResponseRedirect(reverse('historias'))

@login_required
def Operaciones(request):
    citas = SalaOperaciones.objects.all()
    paginator = Paginator(citas, 6)  # mostrar 10 citas por página
    page = request.GET.get('page')
    page_obj = paginator.get_page(page)
    return render(request, 'salaOp/operaciones.html', {'cita': page_obj})

@login_required
def crear_operaciones(request):
    if request.method == 'POST':
        form = SalaForm(request.POST)
        if form.is_valid():
            nueva_cita = form.save(commit=False)
            nueva_cita.save()
            return redirect('operaciones')
    else:
        form = SalaForm()

    context = {'form': form}
    return render(request, 'salaOp/addOperaciones.html', context)

@login_required
def editar_operaciones(request, id):
    # Obtener la cita que se quiere editar
    cita = get_object_or_404(SalaOperaciones, pk=id)

    if request.method == 'POST':
        # Si se envió el formulario con datos, procesarlos
        form = SalaForm(request.POST, instance=cita)
        if form.is_valid():
            form.save()
            return redirect('operaciones')
    else:
        # Si se accede a la vista por primera vez, mostrar el formulario con los datos actuales de la cita
        form = SalaForm(instance=cita)

    return render(request, 'SalaOp/editOperaciones.html', {'form': form, 'cita': cita})

@login_required
def delete_Operaciones(request, id):
    if request.method == 'POST':
        cita = get_object_or_404(SalaOperaciones, pk=id)
        cita.delete()
    return HttpResponseRedirect(reverse('operaciones'))

@login_required
def tecnicos(request):
    historias = Tecnicos.objects.all().order_by('name')
    paginator = Paginator(historias, 6)  # mostrar 10 citas por página
    page = request.GET.get('page')
    page_obj = paginator.get_page(page)
    return render(request, 'tecnicos/Tecnicos.html', {'medico': page_obj})

@login_required
def addTecnico(request):
    if request.method == 'POST':
        form = TecnicosForm(request.POST)
        if form.is_valid():
            new_Nombre_Medico= form.cleaned_data['name']
            new_LastName_Medico= form.cleaned_data['lastName']
            new_Email=form.cleaned_data['email']
            new_Dni=form.cleaned_data['dni']
            new_Cel=form.cleaned_data['celp']
            new_Tipo=form.cleaned_data['tipo']

            new_Medico =Tecnicos(
                name = new_Nombre_Medico,
                lastName = new_LastName_Medico,
                email = new_Email,
                dni = new_Dni,
                celp = new_Cel,
                tipo = new_Tipo,
            )
            new_Medico.save()
            return render(request,'tecnicos/addTecnicos.html',{
                          'form': TecnicosForm(),
                            'success': True
            })
        else:
            form =TecnicosForm()
        return  render(request,'tecnicos/addTecnicos.html',
                       {'form' : TecnicosForm() })
    else:
        form = TecnicosForm()
        return render(request, 'tecnicos/addtecnicos.html', {'form': form})

@login_required
def editTecnicos(request, id):
    medico = Tecnicos.objects.get(pk=id)
    if request.method == 'POST':
        form = TecnicosForm(request.POST, instance=medico)
        if form.is_valid():
            form.save()
            return render(request, 'tecnicos/editTecnicos.html',{
                'form': form,
                'success': True
            })
        else:
            # en caso de que el formulario no sea válido, renderiza el template con el formulario pre-cargado con los datos del paciente existente
            return render(request, 'tecnicos/editTecnicos.html', {
                'form': form,
                'medico': medico
            })
    else:
        # en caso de una solicitud GET, renderiza el formulario pre-cargado con los datos del paciente existente
        form = TecnicosForm(instance=medico)
        return render(request, 'tecnicos/editTecnicos.html', {'form': form, 'medico': medico})

@login_required
def deleteTecnico(request, id):
    if request.method == 'POST':
        medico = Tecnicos.objects.get(pk=id)
        medico.delete()
    return HttpResponseRedirect(reverse('medicos'))

@login_required
def enfermeros(request):
    medico = Enfermeros.objects.all().order_by('name')
    paginator = Paginator(medico, 6)  # mostrar 10 citas por página
    page = request.GET.get('page')
    page_obj = paginator.get_page(page)
    return render(request, 'enfermeros/enfermeros.html', {'medico': page_obj})

@login_required
def addEnfermero(request):
    if request.method == 'POST':
        form = EnfermerosForm(request.POST)
        if form.is_valid():
            new_Nombre_Medico= form.cleaned_data['name']
            new_LastName_Medico= form.cleaned_data['lastName']
            new_Email=form.cleaned_data['email']
            new_Dni=form.cleaned_data['dni']
            new_Cel=form.cleaned_data['celp']
            new_Tipo=form.cleaned_data['tipo']

            new_Medico =Enfermeros(
                name = new_Nombre_Medico,
                lastName = new_LastName_Medico,
                email = new_Email,
                dni = new_Dni,
                celp = new_Cel,
                tipo = new_Tipo,
            )
            new_Medico.save()
            return render(request,'enfermeros/addEnfermeros.html',{
                          'form': EnfermerosForm(),
                            'success': True
            })
        else:
            form =EnfermerosForm()
        return  render(request,'enfermeros/addEnfermeros.html',
                       {'form' : EnfermerosForm() })
    else:
        form = EnfermerosForm()
        return render(request, 'enfermeros/addEnfermeros.html', {'form': form})

@login_required
def editEnfermero(request, id):
    medico = Enfermeros.objects.get(pk=id)
    if request.method == 'POST':
        form = EnfermerosForm(request.POST, instance=medico)
        if form.is_valid():
            form.save()
            return render(request, 'enfermeros/editEnfermeros.html',{
                'form': form,
                'success': True
            })
        else:
            # en caso de que el formulario no sea válido, renderiza el template con el formulario pre-cargado con los datos del paciente existente
            return render(request, 'enfermeros/editEnfermeros.html', {
                'form': form,
                'medico': medico
            })
    else:
        # en caso de una solicitud GET, renderiza el formulario pre-cargado con los datos del paciente existente
        form = EnfermerosForm(instance=medico)
        return render(request, 'enfermeros/editEnfermeros.html', {'form': form, 'medico': medico})

@login_required
def deleteEnfermeros(request, id):
    if request.method == 'POST':
        medico = Enfermeros.objects.get(pk=id)
        medico.delete()
    return HttpResponseRedirect(reverse('enfermeros'))

"""def seleccionar_seguro(request):
    seguros = SeguroPaciente.objects.all()
    pacientes = None  # inicialmente, no se selecciona ningún seguro

    if request.method == 'POST':
        seguro_id = request.POST.get('seguro')
        if seguro_id:
            pacientes = Pacientes.objects.filter(seguro_id=seguro_id)

    return render(request, 'pacientes/seleccionar_seguro.html', {'seguros': seguros, 'pacientes': pacientes})"""

def seleccionar_seguro(request):
    seguros = SeguroPaciente.objects.all()
    pacientes = None  # inicialmente, no se selecciona ningún seguro
    informacion_adicional = None  # información adicional sobre el seguro y beneficios

    if request.method == 'POST':
        seguro_id = request.POST.get('seguro')
        if seguro_id:
            # Obtener pacientes para el seguro seleccionado
            pacientes = Pacientes.objects.filter(seguro_id=seguro_id)

            # Obtener información adicional del seguro y beneficios
            seguro = get_object_or_404(SeguroPaciente, id=seguro_id)
            beneficios = BeneficiosSeguro.objects.filter(seguro=seguro).select_related('cobertura')
            
            informacion_adicional = {'seguro': seguro,'beneficios': beneficios}

    return render(request, 'pacientes/seleccionar_seguro.html', {'seguros': seguros, 'pacientes': pacientes, 'informacion_adicional': informacion_adicional})

@login_required
def equiposMedicos(request):
    medico = EquipoMedico.objects.all().order_by('nombre')
    paginator = Paginator(medico, 10)  # mostrar 10 citas por página
    page = request.GET.get('page')
    page_obj = paginator.get_page(page)
    return render(request, 'equipoMedico/Emedicos.html', {'medico': page_obj})

@login_required
def crearEquipo(request):
    if request.method == 'POST':
        form = EquipoForm(request.POST)
        if form.is_valid():
            equipo = form.save()
            # Realizar cualquier otra operación necesaria con el equipo creado
            return render(request,'equipoMedico/addEmedico.html',{
                          'form': EquipoForm(),
                            'success': True
            })
    else:
        form = EquipoForm()
    
    return render(request, 'equipoMedico/addEmedico.html', {'form': form})

@login_required
def deleteEquipo(request, id):
    if request.method == 'POST':
        medico = EquipoMedico.objects.get(pk=id)
        medico.delete()
    return HttpResponseRedirect(reverse('asignacion'))

@login_required
def editEquipo(request, id):
    medico = EquipoMedico.objects.get(pk=id)
    if request.method == 'POST':
        form = EquipoForm(request.POST, instance=medico)
        if form.is_valid():
            form.save()
            return render(request, 'equipoMedico/editEmedico.html',{
                'form': form,
                'success': True
            })
        else:
            # en caso de que el formulario no sea válido, renderiza el template con el formulario pre-cargado con los datos del paciente existente
            return render(request, 'equipoMedico/editEmedico.html', {
                'form': form,
                'medico': medico
            })
    else:
        # en caso de una solicitud GET, renderiza el formulario pre-cargado con los datos del paciente existente
        form = EquipoForm(instance=medico)
        return render(request, 'equipoMedico/editEmedico.html', {'form': form, 'medico': medico})
    
@login_required
def recetasMedicas(request):
    historias = Recetas.objects.all().order_by('paciente')
    paginator = Paginator(historias, 6)  # mostrar 10 citas por página
    page = request.GET.get('page')
    page_obj = paginator.get_page(page)
    return render(request, 'recetas/recetas.html', {'historia': page_obj})



@login_required
def RecetaPDFView(request, pk):
    # Obtener la historia
    historia = get_object_or_404(Recetas, pk=pk)

    # Crear un objeto PDF en memoria
    buffer = BytesIO()
    c = canvas.Canvas(buffer, pagesize=letter)
    # Crear el documento PDF con ReportLab
    doc = SimpleDocTemplate(buffer, pagesize=letter)
    styles = getSampleStyleSheet()
    style = styles['Normal']

    # Añadir imagen en la parte superior derecha
    logo_path = 'integrador/static/img/logopdf2.png'
    x = 2.5 * inch
    y = 9 * inch
    width = 4 * inch
    height = 2 * inch
    c.drawImage(logo_path, x, y, width, height)
    

    # Añadir título
    
    c.setFont('Helvetica-Bold', 20)
    c.drawCentredString(4.25*inch, 8.5*inch, 'RECETA MEDICA')

    c.drawString(0, 7.5 * inch, '')

    # Crear tabla con datos de la historia
    data = [
        ['Paciente:', f'{historia.paciente.name} {historia.paciente.lastName}'],
        ['Médico:', f'{historia.medico.name} {historia.medico.lastName}'],
        ['Cita:', f'{historia.cita.fechaProgramada}'],
        ['Medicamento:', f'{historia.medicamento}'],
        ['Dosis:', f'{historia.dosis}'],
        ['Observación:', f'{historia.observacion}'],
    ]
    t = Table(data, colWidths=[2*inch, 4*inch])
    t.setStyle(TableStyle([
        ('FONT', (0, 0), (-1, -1), 'Helvetica'),
        ('FONTSIZE', (0, 0), (-1, -1), 12),
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
        ('BACKGROUND', (0, 0), (-1, 0), '#f0f0f0'),
        ('TEXTCOLOR', (0, 0), (-1, 0), 'black'),
        ('GRID', (0, 0), (-1, -1), 1, 'black'),
        ('VALIGN', (0, 0), (-1, -1), 'MIDDLE')
    ]))
    print(t)
      # Obtener el ancho y alto de la tabla
    width, height = t.wrapOn(c, 0, 0)

    # Centrar la tabla en la página
    x = (letter[0] - width) / 2
    y = 6.5 * inch
    t.drawOn(c, x, y)
   
 

    # Configurar la respuesta HTTP
    response = HttpResponse(content_type='application/pdf')
    response['Content-Disposition'] = f'inline; filename="HistoriaClinica_{historia.id}.pdf"'

     # Crear los datos que deseas codificar en el QR
    qr_data = f'Paciente: {historia.paciente.name} {historia.paciente.lastName}\n' \
              f'Médico: {historia.medico.name} {historia.medico.lastName}\n' \
              f'Medicamento : {historia.medicamento}\n' \
              f'Dosis: {historia.dosis}\n' \
              f'Observación: {historia.observacion}'

    # Crear un objeto QRCode
    qr = qrcode.QRCode(
        version=1,
        error_correction=qrcode.constants.ERROR_CORRECT_L,
        box_size=10,
        border=4,
    )

    # Agregar los datos al QRCode
    qr.add_data(qr_data)
    qr.make(fit=True)

    # Crear una imagen QRCode
    qr_image = qr.make_image(fill_color="black", back_color="white")
    qr_buffer = BytesIO()
    qr_image.save(qr_buffer, "PNG")

    # Crear un objeto ImageReader para agregar la imagen QRCode al PDF
    qr_image_reader = ImageReader(qr_buffer)
    
    # Dibujar el código QR en el PDF
    c.drawImage(qr_image_reader, 0.5*inch, 0.5*inch, width=1.5*inch, height=1.5*inch)
    logo_path2 = 'integrador/static/img/QR.png'
    x = 6.5 * inch
    y = 0.5 * inch
    width = 1.5 * inch
    height = 1.5* inch
    c.drawImage(logo_path2, x, y, width, height)
    c.save()
    # Adjuntar el PDF generado a la respuesta HTTP
    response.write(buffer.getvalue())
    buffer.close()

    return response

@login_required
def generar_receta(request):
    if request.method == 'POST':
        form = RecetaForm(request.POST)
        if form.is_valid():
            ticket = form.save()
            # Redireccionar al usuario a la página de confirmación de ticket
            return redirect('recetas')
    else:
        form = RecetaForm()
    return render(request, 'recetas/addRecetas.html', {'form': form})

@login_required
def editar_recetas(request, id):
    # Obtener la cita que se quiere editar
    historia = get_object_or_404(Recetas, pk=id)

    if request.method == 'POST':
        # Si se envió el formulario con datos, procesarlos
        form = RecetaForm(request.POST, instance=historia)
        if form.is_valid():
            form.save()
            return redirect('recetas')
    else:
        # Si se accede a la vista por primera vez, mostrar el formulario con los datos actuales de la cita
        form = RecetaForm(instance=historia)

    return render(request, 'recetas/editReceta.html', {'form': form, 'historia': historia})

@login_required
def delete_recetas(request, id):
    if request.method == 'POST':
        historia = get_object_or_404(Recetas, pk=id)
        historia.delete()
    return HttpResponseRedirect(reverse('recetas'))

@login_required
def examenList(request):
    pacientes = Examen.objects.all().order_by('nombre')
    paginator = Paginator(pacientes, 6)  # muestra 10 pacientes por página
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)
    context = {
        'paciente': page_obj
    }
    return render(request,'examen/examen.html',context)
    
    
@login_required
def createExamen(request):
    if request.method == 'POST':
        form = ExamenForm(request.POST)
        if form.is_valid():
            new_Nombre_Especialidad= form.cleaned_data['nombre']

            new_Especialidad =Examen(
                nombre = new_Nombre_Especialidad,
            )
            new_Especialidad.save()
            return render(request,'examen/addExamen.html',{
                          'form': ExamenForm(),
                            'success': True
            })
        else:
            form =ExamenForm()
        return  render(request,'examen/addExamen.html',
                       {'form' : ExamenForm() })
    else:
        form = ExamenForm()
        return render(request, 'examen/addExamen.html', {'form': form})
    
@login_required
def editarExamen(request,id):
    especialidaddes = Examen.objects.get(pk=id)
    if request.method == 'POST':
        form = ExamenForm(request.POST, instance=especialidaddes)
        if form.is_valid():
            form.save()
            return render(request, 'examen/editExamen.html',{
                'form': form,
                'success': True
            })
        else:
            # en caso de que el formulario no sea válido, renderiza el template con el formulario pre-cargado con los datos del paciente existente
            return render(request, 'examen/editExamen.html', {
                'form': form,
                'paciente': especialidaddes
            })
    else:
        # en caso de una solicitud GET, renderiza el formulario pre-cargado con los datos del paciente existente
        form = EspecialidadForm(instance=especialidaddes)
        return render(request, 'examen/editExamen.html', {'form': form, 'paciente': especialidaddes})
  
@login_required    
def delete_Examen(request,id):
    if request.method == 'POST':
        especialidades2 = Examen.objects.get(pk=id)
        especialidades2.delete()
    return HttpResponseRedirect(reverse('examen'))

@login_required
def horario_medico_list(request):
    pacientes = HorarioMedico.objects.all()
    paginator = Paginator(pacientes, 6)  # muestra 10 pacientes por página
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)
    context = {
        'paciente': page_obj
    }
    return render(request,'horario/horario.html',context)


def horario_medico_create(request):
    if request.method == 'POST':
        form = HorarioMedicoForm(request.POST)
        if form.is_valid():
            new_Nombre_Especialidad= form.cleaned_data['nombre']

            new_Especialidad =HorarioMedico(
                nombre = new_Nombre_Especialidad,
            )
            new_Especialidad.save()
            return render(request,'examen/addExamen.html',{
                          'form': HorarioMedicoForm(),
                            'success': True
            })
        else:
            form =HorarioMedicoForm()
        return  render(request,'horario/addHorario.html',
                       {'form' : HorarioMedicoForm() })
    else:
        form = HorarioMedicoForm()
        return render(request, 'horario/addHorario.html', {'form': form})
    
def horario_medico_update(request, pk):
    especialidaddes = HorarioMedico.objects.get(pk=id)
    if request.method == 'POST':
        form = HorarioMedicoForm(request.POST, instance=especialidaddes)
        if form.is_valid():
            form.save()
            return render(request, 'horario/editHorario.html',{
                'form': form,
                'success': True
            })
        else:
            # en caso de que el formulario no sea válido, renderiza el template con el formulario pre-cargado con los datos del paciente existente
            return render(request, 'horario/editHorario.html', {
                'form': form,
                'paciente': especialidaddes
            })
    else:
        # en caso de una solicitud GET, renderiza el formulario pre-cargado con los datos del paciente existente
        form = EspecialidadForm(instance=especialidaddes)
        return render(request, 'horario/editHorario.html', {'form': form, 'paciente': especialidaddes})

def horario_medico_delete(request, pk):
    horario = get_object_or_404(HorarioMedico, pk=pk)
    if request.method == 'POST':
        horario.delete()
    return redirect('horario')
          